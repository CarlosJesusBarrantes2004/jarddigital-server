from django.db import transaction
from django.db.models import Prefetch
import calendar
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

from decimal import Decimal, ROUND_HALF_UP
from datetime import date, timedelta

# Importamos los modelos y selectores necesarios
from apps.users.models import Usuario, PermisoAcceso
from apps.finances.models import HistoricoPlanilla, Asistencia
from apps.finances.selectors import (
    contar_ventas_instaladas_mes_actual,
    obtener_ventas_pagadas_mes_anterior,
    resumir_pozo_comisiones,
    obtener_regla_comision_vigente,
    contar_inasistencias_mes
)


def generar_excel_asistencias_mensual(queryset_filtrado, mes: int, anio: int) -> HttpResponse:
    """
    Genera un reporte Excel pivotado de asistencias.
    Filas: Asesores | Columnas: Días del mes.
    Aplica estilos condicionales (Celeste para ASISTIÓ, Rojo para NO ASISTIÓ).
    """
    # 1. Preparar la estructura de datos (Pivote en memoria RAM)
    # Formato: { id_usuario: {'nombre': 'Carlos Santisteban', 'dias': { 1: True, 2: False, ... } } }
    datos_asesores = {}

    # Iteramos el queryset (que ya viene filtrado por la vista con el mes y año correctos)
    for asistencia in queryset_filtrado.select_related('id_usuario'):
        uid = asistencia.id_usuario.id
        if uid not in datos_asesores:
            datos_asesores[uid] = {
                'nombre': asistencia.id_usuario.nombre_completo,
                'dias': {}
            }
        # Guardamos el estado usando el día exacto como llave
        datos_asesores[uid]['dias'][asistencia.fecha.day] = asistencia.asistio

    # 2. Configurar el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Asistencias {mes:02d}-{anio}"

    # Estilos Base
    fill_cabecera = PatternFill(start_color='FF4F81BD', end_color='FF4F81BD', fill_type='solid')
    font_cabecera = Font(color='FFFFFFFF', bold=True)
    borde_delgado = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    alineacion_centrada = Alignment(horizontal='center', vertical='center')
    alineacion_izq = Alignment(horizontal='left', vertical='center')

    # ---> ESTILOS CONDICIONALES SOLICITADOS <---
    font_asistio = Font(color='00B0F0', bold=True)  # Color Celeste
    font_no_asistio = Font(color='FF0000', bold=True)  # Color Rojo

    # 3. Construir Cabeceras Dinámicas
    # Averiguamos cuántos días tiene este mes específico
    _, dias_del_mes = calendar.monthrange(anio, mes)

    cabeceras = ["ASESOR"] + [f"{dia:02d}/{mes:02d}/{anio}" for dia in range(1, dias_del_mes + 1)]
    ws.append(cabeceras)

    # Aplicar estilos a la cabecera
    for cell in ws[1]:
        cell.fill = fill_cabecera
        cell.font = font_cabecera
        cell.alignment = alineacion_centrada
        cell.border = borde_delgado

    # 4. Llenado de Datos (Fila por fila)
    for uid, info in datos_asesores.items():
        fila = [info['nombre']]

        # Recorremos cada día del mes (del 1 al 28, 30 o 31)
        for dia in range(1, dias_del_mes + 1):
            estado = info['dias'].get(dia)  # Será True, False, o None (si no hay llave)

            if estado is True:
                fila.append("ASISTIÓ")
            elif estado is False:
                fila.append("NO ASISTIÓ")
            else:
                fila.append("")  # Vacío para feriados, domingos o días sin registrar

        ws.append(fila)

        # 5. Aplicar estilos a la fila recién creada
        fila_actual = ws[ws.max_row]

        # Estilo para el nombre del asesor (Columna A)
        fila_actual[0].alignment = alineacion_izq
        fila_actual[0].border = borde_delgado

        # Estilo para las columnas de los días (Columna B en adelante)
        for idx in range(1, len(fila_actual)):
            celda = fila_actual[idx]
            celda.alignment = alineacion_centrada
            celda.border = borde_delgado

            # Aplicamos los colores de la fuente según el texto
            if celda.value == "ASISTIÓ":
                celda.font = font_asistio
            elif celda.value == "NO ASISTIÓ":
                celda.font = font_no_asistio

    # 6. Auto-Ajuste de Columnas
    # Calculamos el nombre más largo para estirar la primera columna
    max_length_asesor = max(
        [len(str(info['nombre'])) for info in datos_asesores.values()] + [10]) if datos_asesores else 20
    ws.column_dimensions['A'].width = max_length_asesor + 5

    # Estiramos las columnas de fechas para que encaje el texto "NO ASISTIÓ"
    for col_idx in range(2, dias_del_mes + 2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 14

    # 7. Retornamos la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Asistencias_{mes:02d}_{anio}.xlsx"'
    wb.save(response)

    return response


@transaction.atomic
def upsert_asistencia_masiva(datos_validados: list[dict], id_sucursal: int, usuario_peticion) -> int:
    """
    Procesamiento masivo en memoria.
    1 Query de lectura + 1 Query de Inserción Masiva + 1 Query de Actualización Masiva.
    """
    if not datos_validados:
        return 0

    # 1. Extraemos las tuplas de búsqueda para aislar los registros afectados
    ids_usuarios = [item['id_usuario'].id for item in datos_validados]
    fechas = [item['fecha'] for item in datos_validados]

    # 2. Traemos todos los registros existentes (incluso los inactivos) de un solo golpe
    registros_existentes = Asistencia.objects.filter(
        id_usuario_id__in=ids_usuarios,
        fecha__in=fechas
    )

    # Mapeamos en RAM: {(id_usuario, fecha): objeto_asistencia}
    mapa_existentes = {(obj.id_usuario_id, obj.fecha): obj for obj in registros_existentes}

    a_crear = []
    a_actualizar = []

    # 3. Clasificación en memoria
    for item in datos_validados:
        uid = item['id_usuario'].id
        fecha = item['fecha']
        nuevo_estado = item.get('asistio')

        clave = (uid, fecha)

        if clave in mapa_existentes:
            obj = mapa_existentes[clave]
            obj.asistio = nuevo_estado
            obj.activo = True  # Punto 3 resuelto: Reactivación forzada
            a_actualizar.append(obj)
        else:
            a_crear.append(Asistencia(
                id_usuario_id=uid,
                id_sucursal_id=id_sucursal,
                fecha=fecha,
                asistio=nuevo_estado,
                activo=True
            ))

    # 4. Ejecución SQL Óptima
    if a_crear:
        Asistencia.objects.bulk_create(a_crear)
    if a_actualizar:
        Asistencia.objects.bulk_update(a_actualizar, fields=['asistio', 'activo'])

    return len(datos_validados)


def generar_excel_planillas_asesores(queryset_filtrado) -> HttpResponse:
    """
    Genera un reporte Excel con las planillas liquidadas de los asesores.
    Utiliza el Snapshot Pattern (modalidad_aplicada y sede_aplicada) para máxima velocidad y fiabilidad histórica.
    """
    # Consulta limpia y veloz (Traemos solo al usuario para el DNI y Nombre)
    planillas = queryset_filtrado.select_related('id_usuario')

    # Nombre dinámico del archivo
    primer_registro = planillas.first()
    if primer_registro:
        mes_reporte = str(primer_registro.mes_fiscal).zfill(2)
        anio_reporte = str(primer_registro.anio_fiscal)
        nombre_archivo = f"Planilla_Comisiones_{mes_reporte}_{anio_reporte}.xlsx"
    else:
        nombre_archivo = "Planilla_Comisiones_Vacia.xlsx"

    # Preparación del Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilla Comisiones"

    # Estilos
    fill_cabecera = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    font_cabecera = Font(color='FFFFFFFF', italic=True, bold=True)
    borde_delgado = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    alineacion_centrada = Alignment(horizontal='center', vertical='center')
    formato_moneda = '#,##0.00'

    # Cabeceras actualizadas
    cabeceras = ["ITEM", "SEDE - MODALIDAD", "DNI", "ASESOR", "SUELDO (CON DCTO)", "COMISIÓN", "TOTAL"]
    ws.append(cabeceras)

    # Aplicar estilos a la cabecera
    for cell in ws[1]:
        cell.fill = fill_cabecera
        cell.font = font_cabecera
        cell.alignment = alineacion_centrada
        cell.border = borde_delgado

    # Llenado de Datos
    for idx, planilla in enumerate(planillas, start=1):
        usuario = planilla.id_usuario

        # ---> LA SOLUCIÓN ELEGANTE: Snapshot Pattern <---
        modalidad = planilla.modalidad_aplicada or ""
        sede = planilla.sede_aplicada or "SIN SEDE"
        modalidad_sede = f"{sede} - {modalidad}" if modalidad else sede

        dni = usuario.dni if (usuario and hasattr(usuario, 'dni') and usuario.dni) else "SIN DNI"
        nombre = usuario.nombre_completo if usuario else "Desconocido"

        sueldo_base = planilla.sueldo_base_aplicado or Decimal('0.00')
        descuento = planilla.descuento_inasistencias or Decimal('0.00')
        sueldo_descontado = sueldo_base - descuento

        comision = planilla.comision_neta_ganada or Decimal('0.00')
        total = planilla.sueldo_neto_final or Decimal('0.00')

        fila = [
            idx,
            modalidad_sede,
            dni,
            nombre,
            sueldo_descontado,
            comision,
            total
        ]
        ws.append(fila)

        # Aplicar estilos a la fila
        fila_actual = ws[ws.max_row]
        for idx_celda, celda in enumerate(fila_actual):
            celda.alignment = alineacion_centrada
            celda.border = borde_delgado
            if idx_celda in [4, 5, 6]:
                celda.number_format = formato_moneda

    # Auto-Ajuste de Columnas
    max_col_letra = get_column_letter(len(cabeceras))
    ws.auto_filter.ref = f"A1:{max_col_letra}{ws.max_row}"

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = (max_length + 4) if (max_length + 4) >= 12 else 12

    # Retornamos la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)

    return response



# ==========================================
# 1. MOTOR DE PROYECCIÓN (Cálculo individual)
# ==========================================
def proyectar_comisiones_asesor(usuario: Usuario, mes: int, anio: int) -> dict:
    fecha_evaluacion = date(anio, mes, 1)

    # ============================================================
    # MODALIDAD ESTRICTA OPTIMIZADA (0 Consultas N+1)
    # ============================================================
    permiso_activo = None

    if hasattr(usuario, 'permisos_activos_prefetched'):
        if usuario.permisos_activos_prefetched:
            permiso_activo = usuario.permisos_activos_prefetched[0]
    else:
        permiso_activo = PermisoAcceso.objects.filter(
            id_usuario=usuario
        ).select_related(
            'id_modalidad_sede__id_modalidad'
        ).first()

    if not permiso_activo or not getattr(permiso_activo, 'id_modalidad_sede', None):
        raise ValueError(
            f"El asesor {usuario.nombre_completo} no tiene una sede/modalidad activa asignada en el sistema."
        )

    codigo_modalidad = permiso_activo.id_modalidad_sede.id_modalidad.codigo

    nombre_sede = permiso_activo.id_modalidad_sede.nombre \
        if hasattr(permiso_activo.id_modalidad_sede, 'nombre') \
        else str(permiso_activo.id_modalidad_sede)

    # ============================================================
    # ESCENARIO MES ACTUAL → Define el SUELDO BASE
    # ============================================================
    instaladas_mes_actual = contar_ventas_instaladas_mes_actual(usuario.id, mes, anio)
    escenario_sueldo = 'ELITE' if instaladas_mes_actual >= 20 else 'ESTANDAR'

    # ============================================================
    # ESCENARIO MES ANTERIOR → Define las COMISIONES
    # ============================================================
    fecha_mes_anterior = fecha_evaluacion.replace(day=1) - timedelta(days=1)
    mes_anterior = fecha_mes_anterior.month
    anio_anterior = fecha_mes_anterior.year

    instaladas_mes_anterior = contar_ventas_instaladas_mes_actual(
        usuario.id, mes_anterior, anio_anterior
    )
    escenario_comisiones = 'ELITE' if instaladas_mes_anterior >= 20 else 'ESTANDAR'

    # ============================================================
    # REGLAS: Cada escenario busca su propia ReglaComision
    # ============================================================
    # FIX: Se inyectó 'codigo_modalidad' en ambos llamados
    regla_sueldo = obtener_regla_comision_vigente(escenario_sueldo, codigo_modalidad, fecha_evaluacion)
    if not regla_sueldo:
        raise ValueError(
            f"No existe Regla de Comisión para escenario {escenario_sueldo} ({codigo_modalidad}) en {mes}/{anio}."
        )

    regla_comisiones = obtener_regla_comision_vigente(escenario_comisiones, codigo_modalidad, fecha_evaluacion)
    if not regla_comisiones:
        raise ValueError(
            f"No existe Regla de Comisión para escenario {escenario_comisiones} ({codigo_modalidad}) en {mes}/{anio}."
        )

    # ============================================================
    # SUELDO BASE → Depende del escenario del MES ACTUAL
    # ============================================================
    if escenario_sueldo == 'ELITE':
        sueldo_base = regla_sueldo.sueldo_base_elite
    else:
        perfil = getattr(usuario, 'perfil_laboral', None)
        if perfil is None or perfil.sueldo_base_part_time is None:
            raise ValueError(
                f"El asesor {usuario.nombre_completo} no tiene un Perfil Laboral con sueldo base configurado."
            )
        sueldo_base = perfil.sueldo_base_part_time

    # ============================================================
    # COMISIONES → Dependen del escenario del MES ANTERIOR
    # ============================================================
    ventas_pagadas_qs = obtener_ventas_pagadas_mes_anterior(usuario.id, mes, anio)
    resumen = resumir_pozo_comisiones(ventas_pagadas_qs)

    ventas_pagadas = resumen['total_pagadas']
    ventas_av = resumen['total_alto_valor']
    pozo_bruto = Decimal(str(resumen['dinero_bruto']))

    # % del pozo según regla del MES ANTERIOR
    porcentaje_pozo = Decimal('0.00')
    if ventas_pagadas >= regla_comisiones.min_ventas_pagadas_optimo:
        porcentaje_pozo = Decimal('1.00')
    elif ventas_pagadas >= regla_comisiones.min_ventas_pagadas_medio:
        porcentaje_pozo = Decimal('0.50')

    # Multiplicador AV según regla del MES ANTERIOR
    if ventas_av >= regla_comisiones.alto_valor_nivel_3:
        multiplicador_av = Decimal('1.10')
    elif ventas_av >= regla_comisiones.alto_valor_nivel_2:
        multiplicador_av = Decimal('1.00')
    else:
        multiplicador_av = Decimal('0.90')

    # Cálculo final
    comision_neta = (pozo_bruto * porcentaje_pozo * multiplicador_av).quantize(
        Decimal('0.01'), rounding=ROUND_HALF_UP
    )

    # Inasistencias
    faltas = contar_inasistencias_mes(usuario.id, mes, anio)
    descuento_inasistencia = ((sueldo_base / Decimal('30')) * faltas).quantize(
        Decimal('0.01'), rounding=ROUND_HALF_UP
    )

    sueldo_neto_final = (sueldo_base + comision_neta - descuento_inasistencia).quantize(
        Decimal('0.01'), rounding=ROUND_HALF_UP
    )

    return {
        "id_usuario": usuario.id,
        "nombre_completo": usuario.nombre_completo,
        "modalidad_aplicada": codigo_modalidad, # Nuevo campo agregado
        "sede_aplicada": nombre_sede,
        "escenario_sueldo": escenario_sueldo,
        "escenario_comisiones": escenario_comisiones,
        "ventas_instaladas": instaladas_mes_actual,
        "ventas_pagadas": ventas_pagadas,
        "ventas_alto_valor": ventas_av,
        "sueldo_base_aplicado": sueldo_base,
        "pozo_bruto": pozo_bruto,
        "porcentaje_pozo": porcentaje_pozo,
        "multiplicador_av": multiplicador_av,
        "comision_neta": comision_neta,
        "dias_falta": faltas,
        "descuento_faltas": descuento_inasistencia,
        "sueldo_neto_final": sueldo_neto_final
    }

# ==========================================
# 2. MOTOR DE LIQUIDACIÓN MASIVA (RRHH)
# ==========================================
def liquidar_planilla_mensual(mes: int, anio: int, usuario_rrhh: Usuario) -> dict:
    asesores = Usuario.objects.select_related(
        'perfil_laboral'
    ).prefetch_related(
        Prefetch(
            'permisos',
            # FIX: Se retiró el .filter(activo=True).
            # Solo traemos toda la cadena de relaciones sin filtrar por estado.
            queryset=PermisoAcceso.objects.select_related(
                'id_modalidad_sede__id_modalidad'
            ),
            # Guardamos el resultado en un atributo virtual llamado 'permisos_activos_prefetched'
            to_attr='permisos_activos_prefetched'
        )
    ).filter(
        id_rol__codigo='ASESOR',
        activo=True
    )

    creados = 0
    actualizados = 0
    errores = []

    # Transacción maestra
    with transaction.atomic():
        for asesor in asesores:
            try:
                # FIX 1: Savepoint por cada asesor. Si uno explota, los demás sobreviven.
                with transaction.atomic():
                    proyeccion = proyectar_comisiones_asesor(asesor, mes, anio)

                    # FIX 7: Desempacamos la tupla para saber si se creó o actualizó
                    _, created = HistoricoPlanilla.objects.update_or_create(
                        id_usuario=asesor,
                        mes_fiscal=mes,
                        anio_fiscal=anio,
                        defaults={
                            'modalidad_aplicada': proyeccion['modalidad_aplicada'],
                            'sede_aplicada': proyeccion['sede_aplicada'],
                            'ventas_instaladas_mes_actual': proyeccion['ventas_instaladas'],
                            'ventas_pagadas_mes_anterior': proyeccion['ventas_pagadas'],
                            'ventas_alto_valor_pagadas': proyeccion['ventas_alto_valor'],
                            'cantidad_faltas': proyeccion['dias_falta'],

                            'sueldo_base_aplicado': proyeccion['sueldo_base_aplicado'],
                            # FIX 2: Guardamos el factor matemático puro (0.50, 0.90, etc.)
                            'porcentaje_pozo_aplicado': proyeccion['porcentaje_pozo'],
                            'multiplicador_alto_valor': proyeccion['multiplicador_av'],

                            'pozo_comisiones_bruto': proyeccion['pozo_bruto'],
                            'comision_neta_ganada': proyeccion['comision_neta'],
                            'descuento_inasistencias': proyeccion['descuento_faltas'],
                            'sueldo_neto_final': proyeccion['sueldo_neto_final'],

                            'procesado_por': usuario_rrhh
                        }
                    )

                    if created:
                        creados += 1
                    else:
                        actualizados += 1

            except ValueError as e:
                errores.append(f"- {asesor.nombre_completo}: {str(e)}")
            except Exception as e:
                errores.append(f"- {asesor.nombre_completo} (Error DB): {str(e)}")

        # FIX 5: Si hubo errores, lanzamos TODO el reporte para RRHH
        if errores:
            mensaje_error = "La liquidación masiva fue abortada. Corrija los siguientes errores:\n" + "\n".join(errores)
            raise ValueError(mensaje_error)

    # Si todo salió bien
    total = creados + actualizados
    return {
        "mensaje": f"Liquidación exitosa: {creados} planillas nuevas creadas y {actualizados} actualizadas.",
        "total_procesados": total,
        "creados": creados,
        "actualizados": actualizados
    }