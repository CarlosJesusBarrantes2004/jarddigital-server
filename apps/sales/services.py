import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db import transaction
from django.utils import timezone
from django.db.models import Q
from apps.users.models import PermisoAcceso
from rest_framework.exceptions import ValidationError
from .models import Venta, AudioVenta, EstadoAudio, EstadoSOT, SupervisorAsignacion, HistorialAgendaSOT
from .selectors import obtener_ventas_permitidas


def generar_excel_ventas(fecha_inicio: str = None, fecha_fin: str = None, estado_filtro: str = None, usuario_peticion=None) -> HttpResponse:
    """
    Genera el reporte de Excel Multi-hoja con estilos y reglas de negocio.
    """
    # ---> 1. EL CANDADO: Usamos el selector en lugar de objects.all() <---
    ventas_base = obtener_ventas_permitidas(usuario_peticion)

    # ---> 2. OPTIMIZACIÓN EXCEL <---
    ventas_base = ventas_base.select_related(
        'id_distrito_instalacion__id_provincia__id_departamento'
    )

    if fecha_inicio and fecha_fin:
        ventas_base = ventas_base.filter(fecha_venta__range=[fecha_inicio, fecha_fin])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ... [Copia aquí todo tu bloque de definición de estilos: fill_cabecera, bordes, cabeceras, etc.] ...
    fill_cabecera = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    font_cabecera = Font(color='FFFFFFFF', italic=True, bold=True)
    fill_verde = PatternFill(start_color='FF86BF4E', end_color='FF86BF4E', fill_type='solid')
    fill_rojo = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    fill_mes_anio = PatternFill(start_color='FFE4CFC6', end_color='FFE4CFC6', fill_type='solid')
    borde_delgado = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                           bottom=Side(style='thin'))
    alineacion_centrada = Alignment(horizontal='center', vertical='center')

    cabeceras = ["ITEM", "DNI/RUC", "CLIENTE", "SEC", "SOT", "TIPO", "TECN.", "PLAN", "C.FIJO", "SCORE", "COMEN",
                 "EST.SOT", "FECHAINST.", "EST.AUDIO", "AUDIO", "SUBIDA", "SUPERV", "ASESOR", "FECHAVENTA", "MES",
                 "MES INST.", "CELULAR", "C.PAGO", "AÑO", "AÑO INST.", "DEPARTAMENTO", "GÉNERO", "MODALIDAD"]

    hojas_config = {
        'ATENDIDA': ['ATENDIDO', 'ATENDIDA'],
        'EJECUCION': ['EJECUCION'],
        'RECHAZADA': ['RECHAZADO', 'RECHAZADA']
    }

    if estado_filtro:
        estado_filtro = estado_filtro.upper()
        hojas_config = {k: v for k, v in hojas_config.items() if estado_filtro in v or estado_filtro == k}
        if not hojas_config:
            hojas_config = {estado_filtro: [estado_filtro]}

    # ... [Copia aquí todo tu bloque de los FOR loops de ws.append(fila) y estilos de celdas] ...
    for nombre_hoja, codigos_estado in hojas_config.items():
        ventas_hoja = ventas_base.filter(id_estado_sot__codigo__in=codigos_estado)
        ws = wb.create_sheet(title=nombre_hoja)
        ws.append(cabeceras)

        for cell in ws[1]:
            cell.fill = fill_cabecera
            cell.font = font_cabecera
            cell.alignment = alineacion_centrada
            cell.border = borde_delgado

        for idx, venta in enumerate(ventas_hoja, start=1):
            documento = venta.cliente_numero_doc if hasattr(venta,
                                                            'cliente_numero_doc') and venta.cliente_numero_doc else ""
            departamento = ""
            if venta.id_distrito_instalacion and venta.id_distrito_instalacion.id_provincia:
                departamento = venta.id_distrito_instalacion.id_provincia.id_departamento.nombre

            f_venta = venta.fecha_venta
            mes_vta = f_venta.month if f_venta else ""
            anio_vta = f_venta.year if f_venta else ""

            estado_sot_codigo = venta.id_estado_sot.codigo.upper() if venta.id_estado_sot else ""
            f_inst = venta.fecha_rechazo if estado_sot_codigo in ['RECHAZADO', 'RECHAZADA'] else venta.fecha_real_inst

            mes_inst = f_inst.month if f_inst else ""
            anio_inst = f_inst.year if f_inst else ""

            supervisor = ""
            modalidad = ""
            if venta.id_supervisor_vigente:
                if venta.id_supervisor_vigente.id_supervisor: supervisor = venta.id_supervisor_vigente.id_supervisor.nombre_completo
                if venta.id_supervisor_vigente.id_modalidad_sede and venta.id_supervisor_vigente.id_modalidad_sede.id_modalidad: modalidad = venta.id_supervisor_vigente.id_modalidad_sede.id_modalidad.nombre

            genero_inicial = ""
            if venta.cliente_genero:
                genero_mayuscula = venta.cliente_genero.upper()
                genero_inicial = 'M' if genero_mayuscula == 'MASCULINO' else (
                    'F' if genero_mayuscula == 'FEMENINO' else venta.cliente_genero)

            fila = [
                idx, documento, venta.cliente_nombre, venta.codigo_sec or "", venta.codigo_sot or "",
                                                      venta.tipo_venta or "",
                                                      venta.tecnologia or "",
                venta.id_producto.nombre_paquete if venta.id_producto else "",
                venta.id_producto.costo_fijo_plan if venta.id_producto else "",
                                                      venta.score_crediticio or "", venta.comentario_gestion or "",
                venta.id_estado_sot.nombre if venta.id_estado_sot else "",
                f_inst.strftime('%d/%m/%Y') if f_inst else "",
                venta.id_estado_audios.nombre if venta.id_estado_audios else "",
                "✔" if venta.audio_subido else "",
                venta.fecha_subida_audios.strftime('%d/%m/%Y') if venta.fecha_subida_audios else "",
                supervisor, venta.id_asesor.nombre_completo if venta.id_asesor else "",
                f_venta.strftime('%d/%m/%Y') if f_venta else "",
                mes_vta, mes_inst, venta.cliente_telefono if hasattr(venta, 'cliente_telefono') else "", "", anio_vta,
                anio_inst, departamento, genero_inicial, modalidad
            ]
            ws.append(fila)

            fila_actual = ws[ws.max_row]
            for i, celda in enumerate(fila_actual):
                celda.alignment = alineacion_centrada
                celda.border = borde_delgado
                if i == 11 and fila[11]:
                    if fila[11].upper() in ['ATENDIDO', 'ATENDIDA']:
                        celda.fill = fill_verde
                    elif fila[11].upper() in ['RECHAZADO', 'RECHAZADA']:
                        celda.fill = fill_rojo
                elif i == 13 and fila[13] and fila[13].upper() == 'CONFORME':
                    celda.fill = fill_verde
                elif i in [19, 20, 23, 24]:
                    celda.fill = fill_mes_anio

        max_col_letra = get_column_letter(len(cabeceras))
        ws.auto_filter.ref = f"A1:{max_col_letra}{ws.max_row}"
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = (max_length + 5) if (max_length + 5) >= 10 else 10

    if not wb.sheetnames: wb.create_sheet(title="Sin Datos")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Ventas_Clasificado.xlsx"'
    wb.save(response)
    return response



def crear_venta(*, datos_validados: dict, usuario_peticion) -> Venta:
    audios_data = datos_validados.pop('audios', [])

    # 1. Cálculo de Tipo de Venta
    tipo_doc = datos_validados.get('id_tipo_documento')
    if tipo_doc:
        datos_validados['tipo_venta'] = 'CORPORATIVO' if tipo_doc.codigo.upper() == 'RUC' else 'MASIVO'

    datos_validados['id_asesor'] = usuario_peticion
    datos_validados['usuario_creacion'] = usuario_peticion

    # 2. Asignación de Sede y Supervisor
    permiso_sede = PermisoAcceso.objects.filter(
        id_usuario=usuario_peticion, id_modalidad_sede__activo=True
    ).select_related('id_modalidad_sede').first()

    if not permiso_sede:
        raise ValidationError({"error": "No tienes Sede asignada."})

    datos_validados['id_origen_venta'] = permiso_sede.id_modalidad_sede

    hoy = timezone.now().date()
    supervisor_activo = SupervisorAsignacion.objects.filter(
        Q(fecha_fin__isnull=True) | Q(fecha_fin__gte=hoy),
        id_modalidad_sede=permiso_sede.id_modalidad_sede,
        activo=True
    ).first()

    if not supervisor_activo:
        raise ValidationError({"error": "La sede no tiene supervisor activo hoy."})

    datos_validados['id_supervisor_vigente'] = supervisor_activo

    # 3. Limpieza de campos de Backoffice
    campos_backoffice = [
        'codigo_sec', 'codigo_sot', 'fecha_visita_programada', 'bloque_horario', 'id_sub_estado_sot',
        'fecha_real_inst', 'fecha_rechazo', 'comentario_gestion', 'fecha_revision_audios',
        'usuario_revision_audios', 'observacion_audios', 'audio_subido', 'fecha_venta'
    ]
    for campo in campos_backoffice:
        datos_validados.pop(campo, None)

    # 4. Estados por defecto
    datos_validados['id_estado_sot'] = None
    datos_validados['id_estado_audios'] = EstadoAudio.objects.filter(codigo='PENDIENTE').first()
    datos_validados['audio_subido'] = False
    datos_validados['fecha_subida_audios'] = None

    with transaction.atomic():
        venta_creada = Venta.objects.create(**datos_validados)

        # 5. Guardado de Audios Anidados
        audios_a_crear = [AudioVenta(id_venta=venta_creada, **item) for item in audios_data]
        if audios_a_crear:
            AudioVenta.objects.bulk_create(audios_a_crear)

    return venta_creada


def actualizar_venta(*, venta: Venta, datos_validados: dict, usuario_peticion) -> Venta:
    es_backoffice = (usuario_peticion.id_rol and usuario_peticion.id_rol.codigo.upper() == 'BACKOFFICE')

    # Candado de Inmutabilidad
    if venta.id_estado_sot and venta.id_estado_sot.codigo.upper() == 'RECHAZADO' and not es_backoffice:
        raise ValidationError({"error_critico": "Esta venta ha sido RECHAZADA y está cerrada permanentemente."})

    audios_data = datos_validados.pop('audios', None)

    # Recálculo de Tipo de Venta
    nuevo_tipo_doc = datos_validados.get('id_tipo_documento')
    if nuevo_tipo_doc:
        datos_validados['tipo_venta'] = 'CORPORATIVO' if nuevo_tipo_doc.codigo.upper() == 'RUC' else 'MASIVO'

    datos_validados['usuario_modificacion'] = usuario_peticion

    nuevo_codigo_sot = datos_validados.get('codigo_sot')
    nuevo_codigo_sec = datos_validados.get('codigo_sec')
    nuevo_sub_estado = datos_validados.get('id_sub_estado_sot')
    audio_subido_flag = datos_validados.get('audio_subido')
    nueva_fecha_inst = datos_validados.get('fecha_real_inst')
    fecha_visita_antigua = venta.fecha_visita_programada

    with transaction.atomic():
        # 1. Gatillos de Audio
        if audio_subido_flag is True and not venta.audio_subido:
            datos_validados['fecha_subida_audios'] = timezone.now()

        nuevo_estado_audio = datos_validados.get('id_estado_audios')
        if nuevo_estado_audio and nuevo_estado_audio != venta.id_estado_audios:
            datos_validados['usuario_revision_audios'] = usuario_peticion
            datos_validados['fecha_revision_audios'] = timezone.now()

            if nuevo_estado_audio.codigo.upper() == 'RECHAZADO':
                estado_rechazado = EstadoSOT.objects.filter(codigo__iexact='RECHAZADO').first()
                if estado_rechazado:
                    datos_validados['id_estado_sot'] = estado_rechazado
                if not datos_validados.get('fecha_rechazo') and not venta.fecha_rechazo:
                    raise ValidationError({"fecha_rechazo": "Al rechazar por audio, también debe indicar la fecha."})
                if not datos_validados.get('observacion_audios') and not venta.observacion_audios:
                    raise ValidationError({"observacion_audios": "Observación obligatoria al rechazar el audio."})

        # 2. Estampado Fecha de Venta
        if (nuevo_codigo_sot and not venta.codigo_sot) or (nuevo_codigo_sec and not venta.codigo_sec):
            if not venta.fecha_venta:
                datos_validados['fecha_venta'] = timezone.now()

        # 3. Automatizaciones SOT
        if 'id_estado_sot' not in datos_validados:
            if nueva_fecha_inst:
                estado_atendido = EstadoSOT.objects.filter(codigo__iexact='ATENDIDO').first()
                if estado_atendido: datos_validados['id_estado_sot'] = estado_atendido
            elif (nuevo_codigo_sot and not venta.codigo_sot) or (nuevo_codigo_sec and not venta.codigo_sec):
                estado_ejecucion = EstadoSOT.objects.filter(codigo__iexact='EJECUCION').first()
                if estado_ejecucion: datos_validados['id_estado_sot'] = estado_ejecucion

        # 4. Validaciones de Estado Destino
        estado_destino = datos_validados.get('id_estado_sot', venta.id_estado_sot)
        estado_audio_destino = datos_validados.get('id_estado_audios', venta.id_estado_audios)

        if estado_destino:
            if estado_destino.codigo.upper() == 'EJECUCION':
                if not (venta.codigo_sot or nuevo_codigo_sot) or not (venta.codigo_sec or nuevo_codigo_sec):
                    raise ValidationError({"codigo_sot": "Para pasar a EJECUCIÓN, es obligatorio registrar SOT y SEC."})

            elif estado_destino.codigo.upper() == 'ATENDIDO':
                if not estado_audio_destino or estado_audio_destino.codigo.upper() != 'CONFORME':
                    raise ValidationError(
                        {"id_estado_sot": "Bloqueado: No se puede pasar a ATENDIDO si audios no están CONFORME."})
                if not datos_validados.get('fecha_real_inst') and not venta.fecha_real_inst:
                    raise ValidationError({"fecha_real_inst": "Debe ingresar la fecha real de instalación."})

            elif estado_destino.codigo.upper() == 'RECHAZADO':
                if not datos_validados.get('fecha_rechazo') and not venta.fecha_rechazo:
                    raise ValidationError({"fecha_rechazo": "Debe ingresar la fecha de rechazo."})

        if nuevo_sub_estado and (not estado_destino or estado_destino.codigo.upper() != 'EJECUCION'):
            raise ValidationError(
                {"id_sub_estado_sot": "El sub-estado solo se puede asignar si el SOT es 'EJECUCIÓN'."})

        # 5. Guardado Base
        if usuario_peticion.id_rol and usuario_peticion.id_rol.codigo == 'ASESOR':
            datos_validados['solicitud_correccion'] = False
            datos_validados['comentario_gestion'] = None

        for attr, value in datos_validados.items():
            setattr(venta, attr, value)
        venta.save()

        # 6. Actualización Anidada de Audios
        if audios_data is not None:
            audios_existentes = {audio.id: audio for audio in venta.audios.all()}
            for audio_item in audios_data:
                audio_id = audio_item.get('id')
                if audio_id and audio_id in audios_existentes:
                    audio_instance = audios_existentes[audio_id]
                    nueva_url = audio_item.get('url_audio')
                    if nueva_url and nueva_url != audio_instance.url_audio:
                        audio_instance.url_audio = nueva_url
                        audio_instance.conforme = None
                        audio_instance.motivo = None
                        audio_instance.corregido = True

                    audio_instance.nombre_etiqueta = audio_item.get('nombre_etiqueta', audio_instance.nombre_etiqueta)

                    if 'conforme' in audio_item:
                        audio_instance.conforme = audio_item['conforme']
                        audio_instance.corregido = False
                    if 'motivo' in audio_item:
                        audio_instance.motivo = audio_item['motivo']

                    audio_instance.save()
                else:
                    AudioVenta.objects.create(id_venta=venta, **audio_item)

        # 7. Historial de Agenda
        nueva_fecha_visita = datos_validados.get('fecha_visita_programada')
        if nueva_fecha_visita and nueva_fecha_visita != fecha_visita_antigua:
            if nuevo_sub_estado and nuevo_sub_estado.requiere_nueva_fecha:
                HistorialAgendaSOT.objects.create(
                    id_venta=venta, fecha_anterior=fecha_visita_antigua, fecha_nueva=nueva_fecha_visita,
                    id_sub_estado_motivo=nuevo_sub_estado, usuario_responsable=usuario_peticion
                )

    return venta
