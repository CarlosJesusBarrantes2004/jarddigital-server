from rest_framework import viewsets, filters
from rest_framework.permissions import IsAuthenticated
from apps.sales.models import Venta
from apps.sales.serializers import VentaSerializer
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from .filters import VentaFilter

# Importamos utils necesarios para reporte de excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.decorators import action

# Importamos tu papelera de reciclaje y tus aduanas
from apps.core.mixins import SoftDeleteModelViewSet
from apps.users.permissions import SoloLecturaOCrearSiEsJefe
from apps.users.models import PermisoAcceso

from .models import EstadoSOT, SubEstadoSOT, EstadoAudio, Producto, GrabadorAudio
from .serializers import (
    EstadoSOTSerializer,
    SubEstadoSOTSerializer,
    EstadoAudioSerializer,
    ProductoSerializer,
    GrabadorAudioSerializer,
)

from .filters import VentaFilter

from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status

# ==========================================
# 1. CATÁLOGOS Y ESTADOS (Fase 1)
# ==========================================


class EstadoSOTViewSet(SoftDeleteModelViewSet):
    # Usamos order_by('orden') para que el frontend los pinte en el orden lógico del negocio
    queryset = EstadoSOT.objects.all().order_by("orden")
    serializer_class = EstadoSOTSerializer
    permission_classes = [IsAuthenticated, SoloLecturaOCrearSiEsJefe]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ["nombre", "codigo"]


class SubEstadoSOTViewSet(SoftDeleteModelViewSet):
    queryset = SubEstadoSOT.objects.all()
    serializer_class = SubEstadoSOTSerializer
    permission_classes = [IsAuthenticated, SoloLecturaOCrearSiEsJefe]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ["nombre"]


class EstadoAudioViewSet(SoftDeleteModelViewSet):
    queryset = EstadoAudio.objects.all()
    serializer_class = EstadoAudioSerializer
    permission_classes = [IsAuthenticated, SoloLecturaOCrearSiEsJefe]
    filter_backends = [filters.SearchFilter]
    search_fields = ["nombre", "codigo"]


# ==========================================
# 2. OPERATIVOS Y PRODUCTOS (Fase 2)
# ==========================================


class ProductoViewSet(SoftDeleteModelViewSet):
    serializer_class = ProductoSerializer
    permission_classes = [IsAuthenticated, SoloLecturaOCrearSiEsJefe]

    # Activamos los filtros para que el asesor pueda buscar rápido su plan
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = [
        "es_alto_valor",
        "nombre_campana",
        "tipo_solucion",
        "activo",
    ]  # ?es_alto_valor=True
    search_fields = ["nombre_paquete", "nombre_campana"]  # ?search=Max 29.90

    def get_queryset(self):
        activo_param = self.request.query_params.get("activo")
        queryset = Producto.objects.all().order_by("-fecha_inicio_vigencia")
        # Default: solo activos (si no vino el parámetro)
        if activo_param is None:
            queryset = queryset.filter(activo=True)
        return queryset

    @action(detail=True, methods=["patch"], url_path="reactivar")
    def reactivar(self, request, pk=None):
        """
        PATCH /api/sales/productos/{id}/reactivar/
        Reactiva un producto que fue desactivado (soft-delete).
        Usa get_object() con el queryset BASE (todos los registros),
        no el filtrado, por eso puede encontrar inactivos.
        """
        # Buscamos en TODOS los productos, no solo activos
        try:
            producto = Producto.objects.get(pk=pk)
        except Producto.DoesNotExist:
            return Response(
                {"detail": "Producto no encontrado."},
                status=status.HTTP_404_NOT_FOUND,
            )

        producto.activo = True
        producto.save(update_fields=["activo"])

        serializer = self.get_serializer(producto)
        return Response(serializer.data)


class GrabadorAudioViewSet(viewsets.ReadOnlyModelViewSet):
    # El queryset base (trae todos)
    queryset = GrabadorAudio.objects.select_related("id_usuario").all().order_by("id")
    serializer_class = GrabadorAudioSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.SearchFilter]
    filterset_fields = ["activo"]
    search_fields = ["nombre_completo"]

    def get_queryset(self):
        # 1. Obtenemos la lista completa inicial
        queryset = super().get_queryset()

        # 2. Solo aplicamos la lógica de exclusión si estamos LISTANDO (para el dropdown)
        if self.action == "list":
            hoy = timezone.now().date()

            # 3. Buscamos los grabadores que están "Ocupados" hoy.
            # REGLA: Ocupado = Tiene venta creada HOY y la venta está ACTIVA.
            # (Esto bloquea Rechazados, Pendientes, etc. Solo libera si activo=False es decir ANULADA)
            ids_bloqueados = (
                Venta.objects.filter(fecha_creacion__date=hoy, activo=True)
                .exclude(
                    # REGLA DE ORO: El ID 1 (OTROS) NUNCA entra en la lista negra.
                    # Aunque tenga 1000 ventas activas hoy, lo sacamos de la lista de bloqueados.
                    id_grabador_audios=1
                )
                .values_list("id_grabador_audios", flat=True)
            )

            # 4. Excluimos los IDs bloqueados de la respuesta final
            if ids_bloqueados:
                queryset = queryset.exclude(id__in=ids_bloqueados)

        return queryset


# ==========================================
# 3. LA BESTIA: VENTAS (CORE)
# ==========================================


class VentaViewSet(SoftDeleteModelViewSet):
    serializer_class = VentaSerializer
    permission_classes = [IsAuthenticated]  # Todos deben estar logueados

    # Activamos los motores de búsqueda, filtros y ordenamiento
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]

    filterset_class = VentaFilter

    # 2. Buscador libre (Para cuando el cliente llama reclamando y solo dan su DNI)
    search_fields = [
        "cliente_numero_doc",
        "cliente_nombre",
        "codigo_sec",
        "codigo_sot",
        "id_asesor__nombre_completo",
    ]

    # 3. Ordenamiento (Por defecto, las ventas más nuevas arriba)
    ordering_fields = ["fecha_venta", "fecha_creacion"]
    ordering = ["-fecha_venta"]

    # REPORTE DE EXCEL
    @action(detail=False, methods=["get"])
    def exportar_excel(self, request):
        # 1. RECIBIR FECHAS Y ESTADOS DEL FRONTEND
        fecha_inicio = request.query_params.get("fecha_inicio")
        fecha_fin = request.query_params.get("fecha_fin")
        estado_filtro = request.query_params.get(
            "estado_sot"
        )  # Para cuando busquen uno específico

        # 2. LA SÚPER CONSULTA BASE
        ventas_base = Venta.objects.all().select_related(
            "id_producto",
            "id_estado_sot",
            "id_estado_audios",
            "id_asesor",
            "id_supervisor_vigente__id_supervisor",
            "id_supervisor_vigente__id_modalidad_sede__id_modalidad",
            "id_distrito_instalacion__id_provincia__id_departamento",
        )

        if fecha_inicio and fecha_fin:
            ventas_base = ventas_base.filter(
                fecha_venta__range=[fecha_inicio, fecha_fin]
            )

        # 3. CONFIGURACIÓN DEL EXCEL MULTI-HOJA
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Borramos la hoja en blanco por defecto que crea openpyxl

        # 4. DEFINIR ESTILOS
        fill_cabecera = PatternFill(
            start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
        )  # Rojo
        font_cabecera = Font(color="FFFFFFFF", italic=True, bold=True)  # Blanco cursiva

        fill_verde = PatternFill(
            start_color="FF86BF4E", end_color="FF86BF4E", fill_type="solid"
        )  # Verde
        fill_rojo = PatternFill(
            start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
        )  # Rojo (Para Rechazos)
        fill_mes_anio = PatternFill(
            start_color="FFE4CFC6", end_color="FFE4CFC6", fill_type="solid"
        )  # Durazno

        borde_delgado = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        alineacion_centrada = Alignment(horizontal="center", vertical="center")

        cabeceras = [
            "ITEM",
            "DNI/RUC",
            "CLIENTE",
            "SEC",
            "SOT",
            "TIPO",
            "TECN.",
            "PLAN",
            "C.FIJO",
            "SCORE",
            "COMEN",
            "EST.SOT",
            "FECHAINST.",
            "EST.AUDIO",
            "AUDIO",
            "SUBIDA",
            "SUPERV",
            "ASESOR",
            "FECHAVENTA",
            "MES",
            "MES INST.",
            "CELULAR",
            "C.PAGO",
            "AÑO",
            "AÑO INST.",
            "DEPARTAMENTO",
            "GÉNERO",
            "MODALIDAD",
        ]

        # =======================================================
        # DICCIONARIO DE HOJAS Y SUS ESTADOS
        # =======================================================
        # Aquí mapeamos el Nombre de la Hoja -> [Códigos exactos en la BD]
        hojas_config = {
            "ATENDIDA": ["ATENDIDO", "ATENDIDA"],
            "EJECUCION": ["EJECUCION"],
            "RECHAZADA": ["RECHAZADO", "RECHAZADA"],
        }

        # Si el frontend mandó un parámetro específico (?estado_sot=ATENDIDO),
        # filtramos nuestro diccionario para generar SOLO esa hoja
        if estado_filtro:
            estado_filtro = estado_filtro.upper()
            hojas_config = {
                k: v
                for k, v in hojas_config.items()
                if estado_filtro in v or estado_filtro == k
            }
            if not hojas_config:  # Por si mandan algo raro, no rompe el código
                hojas_config = {estado_filtro: [estado_filtro]}

        # =======================================================
        # CREACIÓN DINÁMICA DE HOJAS (Iteramos sobre las 3)
        # =======================================================
        for nombre_hoja, codigos_estado in hojas_config.items():

            # Filtramos de la consulta base SOLO las que corresponden a esta hoja
            ventas_hoja = ventas_base.filter(id_estado_sot__codigo__in=codigos_estado)

            # Si no hay ventas para este estado, pasamos al siguiente (opcional: puedes borrar este 'if' si quieres que genere la hoja vacía)
            # if not ventas_hoja.exists():
            #     continue

            ws = wb.create_sheet(title=nombre_hoja)
            ws.append(cabeceras)

            # Pintar cabeceras
            for cell in ws[1]:
                cell.fill = fill_cabecera
                cell.font = font_cabecera
                cell.alignment = alineacion_centrada
                cell.border = borde_delgado

            # 6. LLENAR LOS DATOS DE ESTA HOJA FILA POR FILA
            for idx, venta in enumerate(ventas_hoja, start=1):
                documento = (
                    venta.cliente_numero_doc
                    if hasattr(venta, "cliente_numero_doc") and venta.cliente_numero_doc
                    else ""
                )

                departamento = ""
                if (
                    venta.id_distrito_instalacion
                    and venta.id_distrito_instalacion.id_provincia
                ):
                    departamento = (
                        venta.id_distrito_instalacion.id_provincia.id_departamento.nombre
                    )

                f_venta = venta.fecha_venta
                mes_vta = f_venta.month if f_venta else ""
                anio_vta = f_venta.year if f_venta else ""

                # ---> ¡NUEVO: LÓGICA DE FECHA DE INSTALACIÓN VS RECHAZO! <---
                estado_sot_codigo = (
                    venta.id_estado_sot.codigo.upper() if venta.id_estado_sot else ""
                )

                if estado_sot_codigo in ["RECHAZADO", "RECHAZADA"]:
                    f_inst = (
                        venta.fecha_rechazo
                    )  # Si es rechazo, usurpa la columna de instalación
                else:
                    f_inst = venta.fecha_real_inst  # Si no, va la instalación normal

                mes_inst = f_inst.month if f_inst else ""
                anio_inst = f_inst.year if f_inst else ""

                supervisor = ""
                modalidad = ""
                if venta.id_supervisor_vigente:
                    if venta.id_supervisor_vigente.id_supervisor:
                        supervisor = (
                            venta.id_supervisor_vigente.id_supervisor.nombre_completo
                        )
                    if (
                        venta.id_supervisor_vigente.id_modalidad_sede
                        and venta.id_supervisor_vigente.id_modalidad_sede.id_modalidad
                    ):
                        modalidad = (
                            venta.id_supervisor_vigente.id_modalidad_sede.id_modalidad.nombre
                        )

                genero_inicial = ""
                if venta.cliente_genero:
                    genero_mayuscula = venta.cliente_genero.upper()
                    if genero_mayuscula == "MASCULINO":
                        genero_inicial = "M"
                    elif genero_mayuscula == "FEMENINO":
                        genero_inicial = "F"
                    else:
                        genero_inicial = venta.cliente_genero

                fila = [
                    idx,
                    documento,
                    venta.cliente_nombre,
                    venta.codigo_sec or "",
                    venta.codigo_sot or "",
                    venta.tipo_venta or "",
                    venta.tecnologia or "",
                    venta.id_producto.nombre_paquete if venta.id_producto else "",
                    venta.id_producto.costo_fijo_plan if venta.id_producto else "",
                    venta.score_crediticio or "",
                    venta.comentario_gestion or "",
                    venta.id_estado_sot.nombre if venta.id_estado_sot else "",
                    f_inst.strftime("%d/%m/%Y") if f_inst else "",
                    venta.id_estado_audios.nombre if venta.id_estado_audios else "",
                    "✔" if venta.audio_subido else "",
                    (
                        venta.fecha_subida_audios.strftime("%d/%m/%Y")
                        if venta.fecha_subida_audios
                        else ""
                    ),
                    supervisor,
                    venta.id_asesor.nombre_completo if venta.id_asesor else "",
                    f_venta.strftime("%d/%m/%Y") if f_venta else "",
                    mes_vta,
                    mes_inst,
                    (
                        venta.cliente_telefono
                        if hasattr(venta, "cliente_telefono")
                        else ""
                    ),
                    "",
                    anio_vta,
                    anio_inst,
                    departamento,
                    genero_inicial,
                    modalidad,
                ]
                ws.append(fila)

                # --- Colores, Bordes y Alineación ---
                fila_actual = ws[ws.max_row]
                for i, celda in enumerate(fila_actual):
                    celda.alignment = alineacion_centrada
                    celda.border = borde_delgado

                    # ---> ¡NUEVO: COLORES PARA ESTADO SOT! (Col 12, index 11) <---
                    if i == 11 and fila[11]:
                        if fila[11].upper() in ["ATENDIDO", "ATENDIDA"]:
                            celda.fill = fill_verde
                        elif fila[11].upper() in ["RECHAZADO", "RECHAZADA"]:
                            celda.fill = fill_rojo  # Pinta la celda de la palabra "RECHAZADA" de rojo

                    # Colorear EST.AUDIO
                    elif i == 13 and fila[13] and fila[13].upper() == "CONFORME":
                        celda.fill = fill_verde

                    # Colorear Meses y Años
                    elif i in [19, 20, 23, 24]:
                        celda.fill = fill_mes_anio

            # Autocorrecciones de la HOJA ACTUAL (Filtro y Autoajuste)
            max_col_letra = get_column_letter(len(cabeceras))
            ws.auto_filter.ref = f"A1:{max_col_letra}{ws.max_row}"

            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            longitud_celda = len(str(cell.value))
                            if longitud_celda > max_length:
                                max_length = longitud_celda
                    except:
                        pass
                ancho_ajustado = (max_length + 5) if (max_length + 5) >= 10 else 10
                ws.column_dimensions[col_letter].width = ancho_ajustado

        # Por si no se creó ninguna hoja (ej. base de datos vacía)
        if not wb.sheetnames:
            wb.create_sheet(title="Sin Datos")

        # 8. PREPARAR Y DESPACHAR
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="Reporte_Ventas_Clasificado.xlsx"'
        )
        wb.save(response)

        return response

    def get_queryset(self):
        user = self.request.user

        queryset = Venta.objects.select_related(
            "id_asesor",
            "id_origen_venta__id_sucursal",
            "id_origen_venta__id_modalidad",
            "id_supervisor_vigente__id_supervisor",
            "id_producto",
            "id_tipo_documento",
            "id_distrito_nacimiento",
            "id_distrito_instalacion",
            "id_sub_estado_sot",
            "id_estado_sot",
            "id_grabador_audios",
            "id_estado_audios",
            "usuario_revision_audios",
            "venta_origen",
        ).all()

        if hasattr(user, "id_rol") and user.id_rol:
            # Convertimos a mayúsculas por seguridad (ej. evita fallos si alguien escribe "Asesor")
            codigo_rol = user.id_rol.codigo.upper()

            # Candado ASESOR: Solo ve sus propias ventas
            if codigo_rol == "ASESOR":
                queryset = queryset.filter(id_asesor=user)

            # Candado SUPERVISOR: Solo ve las ventas de sus sedes asignadas
            elif codigo_rol == "SUPERVISOR":
                sedes_supervisadas = user.asignaciones_supervisor.filter(
                    activo=True, fecha_fin__isnull=True
                ).values_list("id_modalidad_sede", flat=True)
                queryset = queryset.filter(id_origen_venta__in=sedes_supervisadas)

            # Candado BACKOFFICE: Solo ve las ventas de las sedes donde tiene permiso de acceso
            elif codigo_rol == "BACKOFFICE":
                sedes_asignadas = PermisoAcceso.objects.filter(
                    id_usuario=user, id_modalidad_sede__activo=True
                ).values_list("id_modalidad_sede", flat=True)

                queryset = queryset.filter(id_origen_venta__in=sedes_asignadas)

            # Si el rol es DUEÑO, los IFs lo ignoran y se le devuelve el queryset completo.

        return queryset
