from rest_framework.exceptions import ValidationError
from datetime import timedelta
from dateutil.relativedelta import relativedelta
from .models import Seguimiento, SeguimientoMensual

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .selectors import obtener_seguimientos_optimizados


def generar_excel_seguimiento_pendientes(usuario_peticion) -> HttpResponse:
    """
    Genera un reporte Excel con los Seguimientos cuyo Mes 1 NO ha sido pagado.
    Respeta el Row Level Security (RLS) del usuario que lo solicita.
    """
    # 1. Obtenemos el QuerySet base (Ya viene optimizado y con RLS aplicado)
    seguimientos_base = obtener_seguimientos_optimizados(usuario_peticion)

    # 2. EL FILTRO MAGISTRAL: Solo Seguimientos donde el Mes 1 NO esté pagado
    seguimientos_pendientes = seguimientos_base.filter(
        meses_evaluados__mes_numero=1,
        meses_evaluados__pago_cliente_realizado=False
    ).distinct()  # Distinct por si acaso, aunque por arquitectura solo hay un Mes 1 por seguimiento

    # 3. Preparación del Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pendientes Mes 1"

    # Estilos
    fill_cabecera = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    font_cabecera = Font(color='FFFFFFFF', italic=True, bold=True)
    borde_delgado = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    alineacion_centrada = Alignment(horizontal='center', vertical='center')

    cabeceras = ["ID", "DNI/RUC", "CLIENTE", "PLAN", "CODIGO", "ASESOR", "FECHA", "ESTADO"]
    ws.append(cabeceras)

    # Aplicar estilos a la cabecera
    for cell in ws[1]:
        cell.fill = fill_cabecera
        cell.font = font_cabecera
        cell.alignment = alineacion_centrada
        cell.border = borde_delgado

    # 4. Llenado de Datos
    for idx, seguimiento in enumerate(seguimientos_pendientes, start=1):
        venta = seguimiento.id_venta

        # Buscamos el Mes 1 en la memoria (gracias al prefetch_related, esto no hace consultas SQL)
        mes_1 = next((m for m in seguimiento.meses_evaluados.all() if m.mes_numero == 1), None)

        # Formateamos la fecha a "DD/MM" (Ej: "21/04")
        fecha_formateada = ""
        if mes_1 and mes_1.fecha_validacion_pago:
            fecha_formateada = mes_1.fecha_validacion_pago.strftime('%d/%m')

        # Manejo seguro de nulos para la Venta
        dni_ruc = venta.cliente_numero_doc if hasattr(venta, 'cliente_numero_doc') else ""
        cliente = venta.cliente_nombre if hasattr(venta, 'cliente_nombre') else ""
        plan = venta.id_producto.nombre_paquete if (venta and venta.id_producto) else ""
        asesor = venta.id_asesor.nombre_completo if (venta and venta.id_asesor) else ""

        fila = [
            idx,
            dni_ruc,
            cliente,
            plan,
            seguimiento.codigo_pago or "",
            asesor,
            fecha_formateada,
            seguimiento.estado or ""
        ]
        ws.append(fila)

        # Aplicar estilos a la fila recién agregada
        fila_actual = ws[ws.max_row]
        for celda in fila_actual:
            celda.alignment = alineacion_centrada
            celda.border = borde_delgado

    # 5. Auto-Ajuste de Columnas
    max_col_letra = get_column_letter(len(cabeceras))
    ws.auto_filter.ref = f"A1:{max_col_letra}{ws.max_row}"

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = (max_length + 3) if (max_length + 3) >= 12 else 12

    # 6. Retornamos la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Pendientes_Pago_Mes_1.xlsx"'
    wb.save(response)

    return response



def recalcular_fechas_por_nuevo_ciclo(seguimiento: Seguimiento, nuevo_ciclo):
    """
    Recalcula y actualiza los 6 meses existentes en la base de datos
    usando un nuevo ciclo de facturación como base de manera optimizada.
    """
    # Obtenemos los meses ordenados del 1 al 6
    meses = list(seguimiento.meses_evaluados.order_by('mes_numero'))
    if not meses: return

    # --- Mes 1 (Depende directamente del ciclo) ---
    m1 = meses[0]
    m1.fecha_seguimiento = nuevo_ciclo + timedelta(days=10)
    m1.fecha_validacion_pago = nuevo_ciclo + timedelta(days=20)

    # Ya no hacemos m1.save() aquí

    base_fecha_val = m1.fecha_validacion_pago

    # --- Meses 2 al 6 (Regla EOM Iterativa y Autónoma) ---
    for i in range(1, len(meses)):  # Índices 1 al 5 (Meses 2 al 6)
        mes_actual = meses[i]

        # 1. Proyección de Validación amarrada a la base (Mes N - 1)
        nueva_fecha_val = base_fecha_val + relativedelta(months=mes_actual.mes_numero - 1)

        # 2. Proyección de Seguimiento amarrada a la base (Mes N - 2) + 15 días
        # Esto reemplaza a: mes_anterior.fecha_validacion_pago + timedelta(days=15)
        nueva_fecha_seg = base_fecha_val + relativedelta(months=mes_actual.mes_numero - 2) + timedelta(days=15)

        mes_actual.fecha_validacion_pago = nueva_fecha_val
        mes_actual.fecha_seguimiento = nueva_fecha_seg

    # Rematamos con la optimización que hicimos hace un momento
    SeguimientoMensual.objects.bulk_update(
        meses,
        fields=['fecha_seguimiento', 'fecha_validacion_pago']
    )


def actualizar_seguimiento_mensual(*, mes_instance: SeguimientoMensual, datos_validados: dict,
                                   usuario_peticion) -> SeguimientoMensual:
    """
    Servicio encargado de actualizar un registro de Seguimiento Mensual
    aplicando las reglas estrictas de bloqueo de pagos.
    """

    # ---> REGLA: BLOQUEO POR ESTADO DEL PADRE <---
    # Verificamos el estado de la Cabecera (Seguimiento principal)
    if mes_instance.id_seguimiento.estado == 'PENALIZADO':
        raise ValidationError({
            "estado": "Operación denegada. No puedes editar los meses de un seguimiento que se encuentra PENALIZADO. Cambia el estado principal primero."
        })

    nuevo_pago = datos_validados.get('pago_cliente_realizado', mes_instance.pago_cliente_realizado)

    # REGLA 3 y 4: Bloqueo/Desbloqueo de avance por pago
    # Si están intentando marcar o mantener este mes como pagado (True),
    # debemos verificar que el mes INMEDIATAMENTE ANTERIOR también esté pagado.
    if nuevo_pago is True and mes_instance.mes_numero > 1:

        # Buscamos el mes anterior en la base de datos
        mes_anterior = SeguimientoMensual.objects.filter(
            id_seguimiento=mes_instance.id_seguimiento,
            mes_numero=mes_instance.mes_numero - 1
        ).first()

        # Si el mes anterior existe y NO está pagado, levantamos el muro de contención
        if mes_anterior and not mes_anterior.pago_cliente_realizado:
            raise ValidationError({
                "pago_cliente_realizado": f"Regla de Secuencia: No puedes validar el pago del Mes {mes_instance.mes_numero} porque el Mes {mes_anterior.mes_numero} aún no registra pago."
            })

    # REGLA 5: La conformidad no requiere validación extra, ya que es independiente

    # Guardado Base
    for attr, value in datos_validados.items():
        setattr(mes_instance, attr, value)

    # (Opcional) Podemos estampar quién y cuándo modificó esto si tuvieran campos de auditoría

    mes_instance.save()
    return mes_instance